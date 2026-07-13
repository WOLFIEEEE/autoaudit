"""Export an audit result (optionally AI-enriched) to a formatted .xlsx.

Produces three sheets:
  1. Summary        — score, conformance claim, per-level counts.
  2. Issues         — one row per issue with all fields stakeholders ask
                      for (title, description, SC, level, where-to-find,
                      reproduction, recommendation, affected users, …).
  3. WCAG Scorecard — VPAT-style per-SC conformance table.

Formatting:
  - Frozen header row + bold header
  - Severity cells coloured (critical red, serious orange, moderate
    amber, minor grey) so the sheet reads at a glance
  - Autofilter on the Issues sheet
  - Column widths tuned for readability of description / recommendation
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audit.vpat import build_vpat

# Palette calibrated for readability on both Windows + Mac Excel
# defaults. Fills are light so black text stays legible.
_SEVERITY_FILL = {
    "critical": PatternFill("solid", fgColor="F5C2C7"),   # soft red
    "serious":  PatternFill("solid", fgColor="FBD7B0"),   # soft orange
    "moderate": PatternFill("solid", fgColor="FDF0B1"),   # soft amber
    "minor":    PatternFill("solid", fgColor="E2E3E5"),   # soft grey
}
_LEVEL_FILL = {
    "A":   PatternFill("solid", fgColor="C5DBF0"),   # steel blue
    "AA":  PatternFill("solid", fgColor="A9C9E9"),   # stronger blue
    "AAA": PatternFill("solid", fgColor="8FB6DC"),   # strongest blue
}
_CONF_FILL = {
    "Supports":          PatternFill("solid", fgColor="C3E6CB"),
    "Partially Supports": PatternFill("solid", fgColor="FDF0B1"),
    "Does Not Support":   PatternFill("solid", fgColor="F5C2C7"),
    "Not Applicable":     PatternFill("solid", fgColor="E2E3E5"),
    "Not Evaluated":      PatternFill("solid", fgColor="EEEEEE"),
}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="22416B")


def _safe_cell(value: Any) -> Any:
    """Neutralize formulas originating in content from the audited page."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + value
    return value


def _append_safe(ws, values) -> None:
    ws.append([_safe_cell(value) for value in values])

# Column definitions for the Issues sheet. (header, width, accessor)
# Accessor takes an issue dict and returns the cell value.
_ISSUE_COLS: list[tuple[str, int, Any]] = [
    ("#",                 5,  lambda i, idx: idx + 1),
    ("Severity",          10, lambda i, idx: i.get("severity", "")),
    ("Level",             6,  lambda i, idx: i.get("level", "") or ""),
    ("Rule ID",           28, lambda i, idx: i.get("rule", "")),
    ("Module",            14, lambda i, idx: i.get("module", "")),
    ("Principle",         14, lambda i, idx: i.get("principle", "")),
    ("WCAG SC",           14, lambda i, idx: ", ".join(i.get("wcag_criteria") or [])),
    ("Title",             45, lambda i, idx: i.get("title", "")),
    ("Description",       60, lambda i, idx: i.get("description", "")),
    ("Where to find",     40, lambda i, idx: (i.get("ai") or {}).get("location_guide") or _fallback_where(i)),
    ("Reproduction",      40, lambda i, idx: (i.get("ai") or {}).get("reproduction_steps") or ""),
    ("Recommendation",    50, lambda i, idx: (i.get("ai") or {}).get("recommendation") or i.get("fix_suggestion", "")),
    ("Affects users",     30, lambda i, idx: (i.get("ai") or {}).get("user_impact") or _fallback_impact(i)),
    ("Selector",          32, lambda i, idx: (i.get("element") or {}).get("selector", "")),
    ("HTML snippet",      40, lambda i, idx: (i.get("element") or {}).get("html_snippet", "")),
    ("Also detected by",  20, lambda i, idx: ", ".join((i.get("details") or {}).get("also_detected_by") or [])),
]


def _fallback_where(issue: dict[str, Any]) -> str:
    """When AI enrichment is absent, synthesize a usable locator hint
    from the selector + module. Much better than an empty cell and
    still truthful (we only cite facts already in the issue)."""
    el = issue.get("element") or {}
    sel = el.get("selector") or ""
    if sel:
        return f"DOM selector: {sel}"
    return f"module={issue.get('module', '')}"


def _fallback_impact(issue: dict[str, Any]) -> str:
    """Heuristic user-impact line derived from module + severity.

    Low-fidelity by design — when AI enrichment runs, this is replaced
    by the model's per-issue answer. This fallback just keeps the
    spreadsheet informative when no API key was supplied.
    """
    module = (issue.get("module") or "").lower()
    mapping = {
        "screen_reader": "Blind and low-vision users relying on screen readers",
        "keyboard":      "Keyboard-only and motor-impaired users",
        "visual":        "Low-vision and users with reduced motion sensitivity",
        "forms":         "Users filling forms, incl. those with cognitive / motor impairments",
        "media":         "Blind, low-vision, and deaf users consuming media",
        "cognitive":     "Users with cognitive disabilities and low-literacy users",
        "responsive":    "Users on small screens, mobile users, users at 400% zoom",
        "reflow":        "Users at 400% zoom and mobile users in portrait orientation",
        "aria":          "Assistive-tech users (screen readers, switch devices)",
        "widgets":       "Assistive-tech users interacting with custom widgets",
        "dynamic":       "Assistive-tech users after interactive state changes",
        "structure":     "Screen-reader users navigating by headings / landmarks",
        "preferences":   "Users with motion sensitivity or high-contrast mode",
    }
    return mapping.get(module, "Users with disabilities (general)")


# --------------------------------------------------------------------
# Sheet builders


def _autosize_header(ws, header_row: int = 1) -> None:
    for cell in ws[header_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _build_summary_sheet(ws, audit: dict[str, Any]) -> None:
    summary = audit.get("summary") or {}
    ws.title = "Summary"
    rows = [
        ["Field", "Value"],
        ["URL audited", audit.get("url") or ""],
        ["Timestamp", audit.get("timestamp") or ""],
        ["Overall score (0–100)", summary.get("score") or 0],
        ["Grade", summary.get("grade") or ""],
        ["Total issues", summary.get("total_issues") or 0],
        ["", ""],
        ["Severity breakdown", ""],
    ]
    for sev in ("critical", "serious", "moderate", "minor"):
        rows.append([f"  {sev}", (summary.get("by_severity") or {}).get(sev, 0)])
    rows.append(["", ""])
    rows.append(["WCAG conformance", ""])
    for lvl in ("A", "AA", "AAA"):
        n = (summary.get("by_level") or {}).get(lvl, {}).get("issues", 0)
        rows.append([f"  Level {lvl} issues", n])
    conf = summary.get("conformance") or {}
    rows.append(["", ""])
    rows.append(["A conformant?",   "Yes" if conf.get("A_conformant") else "No"])
    rows.append(["AA conformant?",  "Yes" if conf.get("AA_conformant") else "No"])
    rows.append(["AAA conformant?", "Yes" if conf.get("AAA_conformant") else "No"])

    for row in rows:
        _append_safe(ws, row)
    _autosize_header(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def _build_issues_sheet(ws, issues: list[dict[str, Any]]) -> None:
    ws.title = "Issues"
    # Header row.
    _append_safe(ws, [c[0] for c in _ISSUE_COLS])
    _autosize_header(ws)

    # Freeze header + enable autofilter after we know the data range.
    ws.freeze_panes = "A2"

    for idx, issue in enumerate(issues):
        row = [col[2](issue, idx) for col in _ISSUE_COLS]
        _append_safe(ws, row)

        row_num = ws.max_row
        # Severity cell colouring.
        sev = issue.get("severity")
        if sev in _SEVERITY_FILL:
            ws.cell(row=row_num, column=2).fill = _SEVERITY_FILL[sev]
        # Level cell colouring.
        lvl = issue.get("level")
        if lvl in _LEVEL_FILL:
            ws.cell(row=row_num, column=3).fill = _LEVEL_FILL[lvl]
        # Wrap long-text columns so cells don't look truncated at a glance.
        for col_idx in (8, 9, 10, 11, 12, 13):  # title .. affects
            ws.cell(row=row_num, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    # Column widths.
    for i, (_, width, _f) in enumerate(_ISSUE_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Autofilter across the whole issues range so stakeholders can
    # slice by severity / level / module without extra setup.
    if ws.max_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_ISSUE_COLS))}{ws.max_row}"


def _build_vpat_sheet(ws, audit: dict[str, Any], target_level: str) -> None:
    ws.title = "WCAG Scorecard"
    vpat = build_vpat(audit, target_level=target_level)
    _append_safe(ws, [
        f"WCAG 2.2 conformance (target: Level {target_level})",
        "", "", "", "",
    ])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _append_safe(ws, [vpat.get("overall_claim") or ""])
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    _append_safe(ws, [])

    _append_safe(ws, ["SC", "Title", "Level", "Conformance", "Remarks"])
    header_row = ws.max_row
    _autosize_header(ws, header_row=header_row)

    for r in vpat["rows"]:
        _append_safe(
            ws, [r["sc"], r["title"], r["level"], r["conformance"], r["remarks"]]
        )
        fill = _CONF_FILL.get(r["conformance"])
        if fill:
            ws.cell(row=ws.max_row, column=4).fill = fill
        ws.cell(row=ws.max_row, column=5).alignment = Alignment(wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"
    widths = [8, 38, 8, 22, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------
# Public entry points


def build_xlsx_bytes(audit: dict[str, Any], *, target_level: str = "AA") -> bytes:
    """Return the audit as an in-memory .xlsx byte string.

    Useful for HTTP responses (Content-Type:
    application/vnd.openxmlformats-officedocument.spreadsheetml.sheet).
    """
    wb = Workbook()
    # Workbook comes with a default sheet we'll repurpose as Summary.
    _build_summary_sheet(wb.active, audit)
    _build_issues_sheet(wb.create_sheet(), audit.get("issues") or [])
    _build_vpat_sheet(wb.create_sheet(), audit, target_level)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_xlsx(audit: dict[str, Any], path: str, *, target_level: str = "AA") -> str:
    """Write the audit to `path` and return that path.

    Creates parent directories on demand so callers don't have to
    precompute the directory tree — a common trip for CLI users.
    """
    from pathlib import Path

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(build_xlsx_bytes(audit, target_level=target_level))
    return str(p)
