"""Security regressions for spreadsheet export."""

from openpyxl import Workbook

from audit.export_xlsx import _build_issues_sheet


def test_issue_text_cannot_become_excel_formula():
    workbook = Workbook()
    sheet = workbook.active
    issue = {
        "severity": "serious",
        "level": "AA",
        "rule": "test-rule",
        "module": "test",
        "wcag_criteria": ["1.1.1"],
        "principle": "perceivable",
        "title": '=HYPERLINK("https://attacker.invalid", "click")',
        "description": "+cmd|' /C calc'!A0",
        "element": {"selector": "#target"},
    }

    _build_issues_sheet(sheet, [issue])

    title_cell = sheet.cell(row=2, column=8)
    description_cell = sheet.cell(row=2, column=9)
    assert title_cell.data_type != "f"
    assert description_cell.data_type != "f"
    assert title_cell.value.startswith("'=")
    assert description_cell.value.startswith("'+")
